import tkinter
from tkinter import *
import openpyxl
from openpyxl import Workbook
import operator
import pandas as pd
import numpy as np
from playsound import playsound
import datetime
from datetime import datetime
from PIL import Image, ImageTk
import os.path
import cv2

widget = tkinter.Tk()
This_folder = os.path.dirname(os.path.abspath(__file__)) #패스설정
Excel_file = os.path.join(This_folder, 'auction.xlsx') #패스설정
##Greeting = "0"
##JPG_file = os.path.join(This_folder, f'{Greeting}.jpg')
wb = openpyxl.load_workbook(Excel_file)
wb.get_sheet_names()
test = wb.get_sheet_by_name("Sheet1")

# 프로그램 이름
widget.title("Glovis Auction system")
  
# 윈도우 사이즈
widget.geometry("1440x900")

# 글로비스 or 오토벨 로고
image = tkinter.PhotoImage(file="glovis11.png")
tkinter.Label(widget, image = image).place(x = 1300, y = 0)

# SEQ 키인
tkinter.Label(widget, text="SEQ", font=30).grid(row=5, column=0) #그리드 설정
# tkinter.Label(widget, text="SEQ", font=30).place(x = 100, y = 200) #플레이스 설정
SEQ_Input = tkinter.Entry(widget, width=30, justify='center', bg='yellow')
#SEQ_Input.place(x = 50, y = 200)
SEQ_Input.grid(row=5, column=1)

#Pic = ImageTk.PhotoImage(Image.open(f"{Pic1}.jpg"))
#Pic = tkinter.PhotoImage(Image.open("c:/auction_system/"+f"{Pic1}.jpg"))

#차량 사진 불러오기 집작업

#def Updation():                       #configure 에러 모듈
#    path = SEQ_Input.get()+".png"
#    img = tkinter.PhotoImage(file=path)
#    label_Img.configure(image = img)
#    label_Img.image=img

def IMG_fix1():
    txt_s = str(SEQ_Input.get())
    py_Img=tkinter.PhotoImage(file=f"{txt_s}.png")

def IMG_fix():
    txt_s = str(SEQ_Input.get())
    py_Img=tkinter.PhotoImage(file=f"{txt_s}.png")
    label_Img = tkinter.Label(widget, text = "Image Load...", image = py_Img).place(x = 600, y = 300)
    print (txt_s)
    print (SEQ_Input.get())
# label_Img = tkinter.Label(widget, text = "Image Load...", image = py_Img).grid(row=10, column=10)
#label_Img.pack()


# 차량 사진 불러오기
#image1 = ImageTk.PhotoImage(file=f"{Pic1}.jpg")
#tkinter.Label(widget, image = image1).place(x = 600, y = 200)
#image1 = tkinter.PhotoImage(file="c:/auction_system/" + f"{Pic1}.jpg")
#tkinter.Label(widget, image = image1).place(x = 600, y = 200)

#im = Image.open(JPG_file)
#ph = ImageTk.PhotoImage(im)
#label = tkinter.Label(widget, image=ph).place(x = 600,y = 100)

#Greeting = "0"
#JPG_file = os.path.join(This_folder, f'{Greeting}.jpg')

img_src = "1.png"
img = cv2.imread(img_src)

def IMG_show():
    cv2.imshow('Car IMG1', img)

#print (Greeting)

# 버튼 설정
Sq_Button = tkinter.Button(widget, text="Sequencing", width=10, font=30, command=lambda:[Sq1(), Sq2(), Sq3(), Sq4(), Sq5(), IMG_fix1(), IMG_fix()]).grid(row=5, column=3) # 그리드 설정
# Sq_Button = tkinter.Button(widget, text="Sequencing", width=10, font=30, command=lambda:[Sq1(), Sq2(), Sq3(), Sq4(), Sq5(), Updation()]).place(x = 400, y = 200)

# Sequencing
def Sq1():    
    txt1.set(SEQ_Input.get())

def Sq2():    
    txt2.set(test.cell(txt1.get()+1, 3).value)

def Sq3():    
    txt3.set(test.cell(txt1.get()+1, 4).value)

def Sq4():    
    txt4.set(test.cell(txt1.get()+1, 5).value)

def Sq5():    
    txt5.set(test.cell(txt1.get()+1, 6).value)

# 출력 부문 설정
txt1 = tkinter.IntVar()

#def Updation():
 #   if txt1.get() > 0 :
 #       im = Image.open(JPG_file)
 #       ph = ImageTk.PhotoImage(im)
 #       label = tkinter.Label(widget, image=ph).place(x = 600,y = 100)
 #       print (Greeting)
 #       print (JPG_file)
############################################################################
#txt_s = str(SEQ_Input.get()) + "1" #포토이미지 이름 변경
#Pic1 = txt_s
#py_Img=tkinter.PhotoImage(file=f"{Pic1}.png")
#py_Img=tkinter.PhotoImage(file=f"{txt_s}.png") #포토이미지 위치 지정
############################################################################
txt2 = tkinter.StringVar()
txt3 = tkinter.StringVar() #
txt4 = tkinter.StringVar() #
txt5 = tkinter.StringVar() #
txt6 = tkinter.StringVar() #

tkinter.Label(widget, text="                                                            ").grid(row=0, column=0) # 공백
tkinter.Label(widget, text="                                                            ").grid(row=1, column=0) # 공백
tkinter.Label(widget, text="                                                            ").grid(row=2, column=0) # 공백
tkinter.Label(widget, text="                                                            ").grid(row=3, column=0) # 공백
tkinter.Label(widget, text="                                                            ").grid(row=4, column=0) # 공백
tkinter.Label(widget, text="                                                            ").grid(row=6, column=0) # 공백
tkinter.Label(widget, text="                                                            ").grid(row=7, column=0) # 공백

tkinter.Label(widget, text="Squnece", font=30).grid(row=8, column=0) #시퀀스
tkinter.Label(widget, textvariable=txt1, width=30, bg='white', anchor='w').grid(row=8, column=1, padx=5) # 게시용1

tkinter.Label(widget, text="Model", font=30).grid(row=9, column=0) #모델명
tkinter.Label(widget, textvariable=txt2, width=30, bg='white', anchor='w').grid(row=9, column=1, padx=5) # 게시용2

tkinter.Label(widget, text="Regn No.", font=30).grid(row=10, column=0) #등록번호
tkinter.Label(widget, textvariable=txt3, width=30, bg='white', anchor='w').grid(row=10, column=1, padx=5) # 게시용3

tkinter.Label(widget, text="Km", font=30).grid(row=11, column=0) #키로수
tkinter.Label(widget, textvariable=txt4, width=30, bg='white', anchor='w').grid(row=11, column=1, padx=5) # 게시용4

tkinter.Label(widget, text="No. of Owner", font=30).grid(row=12, column=0) #차량 소유자 수
tkinter.Label(widget, textvariable=txt5, width=30, bg='white', anchor='w').grid(row=12, column=1, padx=5) # 게시용4

#tkinter.Label(widget, text="Squnece", font=30).place(x = 400, y = 250) #시퀀스
#tkinter.Label(widget, textvariable=txt1, width=30, bg='white', anchor='w').place(x = 350, y = 250) # 게시용1

#tkinter.Label(widget, text="Model", font=30).place(x = 400, y = 300) #모델명
#tkinter.Label(widget, textvariable=txt2, width=30, bg='white', anchor='w').place(x = 350, y = 300) # 게시용2

#tkinter.Label(widget, text="Regn No.", font=30).place(x = 400, y = 350) #등록번호
#tkinter.Label(widget, textvariable=txt3, width=30, bg='white', anchor='w').place(x = 350, y = 350) # 게시용3

#tkinter.Label(widget, text="Km", font=30).place(x = 400, y = 400) #키로수
#tkinter.Label(widget, textvariable=txt4, width=30, bg='white', anchor='w').place(x = 350, y = 400) # 게시용4

#tkinter.Label(widget, text="No. of Owner", font=30).place(x = 400, y = 450) #차량 소유자 수
#tkinter.Label(widget, textvariable=txt5, width=30, bg='white', anchor='w').place(x = 350, y = 450) # 게시용4


widget.mainloop()