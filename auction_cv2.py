import tkinter
from tkinter import *
import openpyxl
from openpyxl import Workbook
import operator
import pandas as pd
import numpy as np
from playsound import playsound
import datetime
from datetime import datetime
from PIL import Image, ImageTk
import os.path
import cv2
import pyautogui

widget = tkinter.Tk()
widget1 = tkinter.Toplevel()
This_folder = os.path.dirname(os.path.abspath(__file__)) #패스설정
Excel_file = os.path.join(This_folder, 'auction.xlsx') #패스설정
##Greeting = "0"
##JPG_file = os.path.join(This_folder, f'{Greeting}.jpg')
wb = openpyxl.load_workbook(Excel_file)
wb.get_sheet_names()
test = wb.get_sheet_by_name("Sheet1")

# 프로그램 이름
widget.title("Glovis Auction system")
widget1.title("Glovis Auction Picture")

# 윈도우 사이즈
widget.geometry("600x1060+0+0")
widget1.geometry("1320x1060+600+0")

# 글로비스 or 오토벨 로고
image = tkinter.PhotoImage(file="c:/auction_system/glovis11.png")
glovis_logo = tkinter.Label(widget, image = image).place(x = 0, y = 20)
autoimage = tkinter.PhotoImage(file="c:/auction_system/autobell.png")
autobell_logo = tkinter.Label(widget1, image = autoimage).place(x = 850, y = 0)
Bimage = tkinter.PhotoImage(file="c:/auction_system/background.png")
autobell_logo = tkinter.Label(widget1, image = Bimage).place(x = 0, y = 800)
#tkinter.Label(widget, image = image).grid(row=0, column=0)

# SEQ 키인
tkinter.Label(widget, text="SEQ", font=("Malgun Gothic", "20")).place(x=50, y = 175) #그리드 설정
# tkinter.Label(widget, text="SEQ", font=30).place(x = 100, y = 200) #플레이스 설정
SEQ_Input = tkinter.Entry(widget, width=18, justify='center', bg='yellow', font=("Malgun Gothic", "20"))
#SEQ_Input.place(x = 50, y = 200)
SEQ_Input.place(x=230, y = 175)

#Pic = ImageTk.PhotoImage(Image.open(f"{Pic1}.jpg"))
#Pic = tkinter.PhotoImage(Image.open("c:/auction_system/"+f"{Pic1}.jpg"))

#차량 사진 불러오기 집작업

#def Updation():                       #configure 에러 모듈
#    path = SEQ_Input.get()+".png"
#    img = tkinter.PhotoImage(file=path)
#    label_Img.configure(image = img)
#    label_Img.image=img

#def IMG_fix1():
#    txt_s = str(SEQ_Input.get())
#    py_Img=tkinter.PhotoImage(file=f"{txt_s}.png")

#def IMG_fix():
#    txt_s = str(SEQ_Input.get())
#    py_Img=tkinter.PhotoImage(file=f"{txt_s}.png")
#    label_Img = tkinter.Label(widget, text = "Image Load...", image = py_Img).place(x = 600, y = 300)
#    print (txt_s)
#    print (SEQ_Input.get())
# label_Img = tkinter.Label(widget, text = "Image Load...", image = py_Img).grid(row=10, column=10)
#label_Img.pack()


# 차량 사진 불러오기
#image1 = ImageTk.PhotoImage(file=f"{Pic1}.jpg")
#tkinter.Label(widget, image = image1).place(x = 600, y = 200)
#image1 = tkinter.PhotoImage(file="c:/auction_system/" + f"{Pic1}.jpg")
#tkinter.Label(widget, image = image1).place(x = 600, y = 200)

#im = Image.open(JPG_file)
#ph = ImageTk.PhotoImage(im)
#label = tkinter.Label(widget, image=ph).place(x = 600,y = 100)

#Greeting = "0"
#JPG_file = os.path.join(This_folder, f'{Greeting}.jpg')

#######CV 방식 기본 셋팅
#img_src = "1.png"
#img = cv2.imread(img_src)

#def IMG_show():
#    cv2.imshow('Car IMG1', img)
#######

#CV 방식 응용 셋팅

def IMG_show1():                #_0 사진
    imim = str(SEQ_Input.get()) #SEQ 받기
    img_src ="c:/auction_system/" + f"{imim}_0.jpg" #파일이름 설정해주기
    img = cv2.imread(img_src) # 읽을 파일 결정
    img1 = cv2.resize(img, (660, 600))
    winname = "AutoPIC1"
    cv2.namedWindow(winname)  
    cv2.moveWindow(winname, 601, 200)  
    cv2.imshow(winname, img1)

def IMG_show2(): #_1 사진
    imim = str(SEQ_Input.get()) #SEQ 받기
    img_src1 ="c:/auction_system/" + f"{imim}_1.jpg" #파일이름 설정해주기
    img3 = cv2.imread(img_src1) # 읽을 파일 결정
    img4 = cv2.resize(img3, (660, 600))
    winname1 = "AutoPIC2"
    cv2.namedWindow(winname1)   
    cv2.moveWindow(winname1, 1261, 200)   
    cv2.imshow(winname1, img4)

#def IMG_show3(): #_2 사진
#    imim = str(SEQ_Input.get()) #SEQ 받기
#    img_src2 ="c:/auction_system/" + f"{imim}_2.png" #파일이름 설정해주기
#    img5 = cv2.imread(img_src2) # 읽을 파일 결정
#    img6 = cv2.resize(img5, (660, 485))
#    winname2 = "AutoPIC3"
#    cv2.namedWindow(winname2) 
#    cv2.moveWindow(winname2, 601, 516)   
#    cv2.imshow(winname2, img6)

#def IMG_show4(): #_3 사진
#    imim = str(SEQ_Input.get()) #SEQ 받기
#    img_src3 ="c:/auction_system/" + f"{imim}_3.png" #파일이름 설정해주기
#    img7 = cv2.imread(img_src3) # 읽을 파일 결정
#    img8 = cv2.resize(img7, (660, 485))
#    winname3 = "AutoPIC4"
#    cv2.namedWindow(winname3)   
#    cv2.moveWindow(winname3, 1261, 516)   
#    cv2.imshow(winname3, img8)

    #cv2.imshow('Car IMG1', img)
    #cv2.namedWindow(car_img1)
    #cv2.moveWindow(car_img1, 40, 30)

#print (Greeting)

# 버튼 설정
Sq_Button = tkinter.Button(widget, text="Enter", width=5, font=("Malgun Gothic", "15"), command=lambda:[Sq1(), Sq2(), Sq3(), Sq4(), Sq5(), IMG_show2(), IMG_show1(), S_Price(), B_Price(), compare(), tab6()]).place(x=520, y = 170) # 그리드 설정
# Sq_Button = tkinter.Button(widget, text="Seq", width=10, font=("Malgun Gothic", "15"), command=lambda:[Sq1(), Sq2(), Sq3(), Sq4(), Sq5(), IMG_show4(), IMG_show3(), IMG_show2(), IMG_show1(), S_Price(), B_Price(), compare()]).place(x=450, y = 140) # 그리드 설정
# Sq_Button = tkinter.Button(widget, text="Sequencing", width=10, font=30, command=lambda:[Sq1(), Sq2(), Sq3(), Sq4(), Sq5(), Updation()]).place(x = 400, y = 200)
Bid_Button1 = tkinter.Button(widget, text="RS. 1,000", width=10, font=("Malgun Gothic", "15"), command=lambda:[B_1000(), B_Price(), L_cost(), Bidder_Ent(), B_Status(), compare(), Bidders1(), tab7()]).place(x=70, y = 950)
Bid_Button3 = tkinter.Button(widget, text="RS. 3,000", width=10, font=("Malgun Gothic", "15"), command=lambda:[B_3000(), B_Price(), L_cost(), Bidder_Ent(), B_Status(), compare(), Bidders1(), tab7()]).place(x=220, y = 950)
Bid_Button5 = tkinter.Button(widget, text="RS. 5,000", width=10, font=("Malgun Gothic", "15"), command=lambda:[B_5000(), B_Price(), L_cost(), Bidder_Ent(), B_Status(), compare(), Bidders1(), tab7()]).place(x=370, y = 950)
Bid_Button5 = tkinter.Button(widget, text="DONE", width=5, fg = "white", bg = "black", font=("Malgun Gothic", "15"), command=lambda:[Winning(), Winning1(), tab1()]).place(x=520, y = 950)

#비더 표시  
#tkinter.Label(widget, text="Winner", font=("Malgun Gothic", "15")).place(x=100, y = 145)
# tkinter.Label(widget, text="SEQ", font=30).place(x = 100, y = 200) #플레이스 설정
Winner_Input = tkinter.Entry(widget, width=5, justify='center', bg='yellow', font=("Malgun Gothic", "20"))
#SEQ_Input.place(x = 50, y = 200)
Winner_Input.place(x=250, y = 900)


# Sequencing
def Sq1():    
    txt1.set(SEQ_Input.get())

def Sq2():    
    txt2.set(test.cell(txt1.get()+1, 3).value)

def Sq3():    
    txt3.set(test.cell(txt1.get()+1, 4).value)

def Sq4():    
    txt4.set(test.cell(txt1.get()+1, 5).value)

def Sq5():    
    txt5.set(test.cell(txt1.get()+1, 6).value)

def S_Price():    #기본 가격 정보 제시
    StartPrice.set(test.cell(txt1.get()+1, 7).value)

def L_cost():
    LandCost.set(test.cell(txt1.get()+1, 8).value)

def B_Price():    #현재 비드 프라이스를 가지고 온다
    BidPrice.set(test.cell(txt1.get()+1, 9).value)

def B_1000():
    test.cell(txt1.get()+1, 9).value = test.cell(txt1.get()+1, 9).value + 1000
    wb.save("c:/auction_system/auction.xlsx")

def B_3000():
    test.cell(txt1.get()+1, 9).value = test.cell(txt1.get()+1, 9).value + 3000
    wb.save("c:/auction_system/auction.xlsx")

def B_5000():
    test.cell(txt1.get()+1, 9).value = test.cell(txt1.get()+1, 9).value + 5000
    wb.save("c:/auction_system/auction.xlsx")

def Bidder_Ent():
    test.cell(txt1.get()+1, 10).value = Winner_Input.get()
    wb.save("c:/auction_system/auction.xlsx")

def B_Status():
    if LandCost.get() > BidPrice.get() :
        test.cell(txt1.get()+1, 11).value = "Under Negotiation.."
        wb.save("c:/auction_system/auction.xlsx")

    else :
        test.cell(txt1.get()+1, 11).value = "Too close ~!"
        wb.save("c:/auction_system/auction.xlsx")
        playsound("c:/auction_system/OK.wav")

def compare():
    BidStatus.set(test.cell(txt1.get()+1, 11).value)

def Bidders1():
    BiderStatus.set(test.cell(txt1.get()+1, 10).value)

def Winning():
    test.cell(txt1.get()+1, 11).value = "Winner~!"
    wb.save("c:/auction_system/auction.xlsx")
    playsound("c:/auction_system/OK.wav")

def Winning1():
    BidStatus.set(test.cell(txt1.get()+1, 11).value)

def tab7():#오토키보드 함수 설정
    pyautogui.press('tab', presses=7)

def tab1():#오토키보드 함수 설정
    pyautogui.press('tab', presses=1)

def tab6():
    pyautogui.press('tab', presses=6)
    pyautogui.press('delete', presses=1)



def Next():#오토키보드 함수 설정
    pyautogui.press(test.cell(txt1.get()+2, 1).value)
    Sq_Button

def Next_1():
    test.cell(txt1.get()+2, 1).value


# 출력 부문 설정
txt1 = tkinter.IntVar()

#def Updation():
 #   if txt1.get() > 0 :
 #       im = Image.open(JPG_file)
 #       ph = ImageTk.PhotoImage(im)
 #       label = tkinter.Label(widget, image=ph).place(x = 600,y = 100)
 #       print (Greeting)
 #       print (JPG_file)
############################################################################
#txt_s = str(SEQ_Input.get()) + "1" #포토이미지 이름 변경
#Pic1 = txt_s
#py_Img=tkinter.PhotoImage(file=f"{Pic1}.png")
#py_Img=tkinter.PhotoImage(file=f"{txt_s}.png") #포토이미지 위치 지정
############################################################################
txt2 = tkinter.StringVar()
txt3 = tkinter.StringVar() #
txt4 = tkinter.StringVar() #
txt5 = tkinter.StringVar() #
txt6 = tkinter.StringVar()
StartPrice = tkinter.IntVar() #
BidPrice = tkinter.IntVar() #
LandCost = tkinter.IntVar()
BidStatus = tkinter.StringVar()
BiderStatus = tkinter.StringVar()

#tkinter.Label(widget, text="                                                            ").grid(row=0, column=0) # 공백
#tkinter.Label(widget, text="                                                            ").grid(row=1, column=0) # 공백
#tkinter.Label(widget, text="                                                            ").grid(row=2, column=0) # 공백
#tkinter.Label(widget, text="                                                            ").grid(row=3, column=0) # 공백
#tkinter.Label(widget, text="                                                            ").grid(row=4, column=0) # 공백
#tkinter.Label(widget, text="                                                            ").grid(row=6, column=0) # 공백
#tkinter.Label(widget, text="                                                            ").grid(row=7, column=0) # 공백

tkinter.Label(widget, text="Sequence", font=("Malgun Gothic", "20")).place(x=50, y = 220)
tkinter.Label(widget, textvariable=txt1, width=18, justify='center', bg='white', font=("Malgun Gothic", "20")).place(x=230, y = 220) # 게시용1

tkinter.Label(widget, text="Model", font=("Malgun Gothic", "20")).place(x=50, y = 265) #모델명
tkinter.Label(widget, textvariable=txt2, width=18, justify='center', bg='white', font=("Malgun Gothic", "20")).place(x=230, y = 265) # 게시용2

tkinter.Label(widget, text="Regn No.", font=("Malgun Gothic", "20")).place(x=50, y = 310) #등록번호
tkinter.Label(widget, textvariable=txt3, width=18, justify='center', bg='white', font=("Malgun Gothic", "20")).place(x=230, y = 310) # 게시용3

tkinter.Label(widget, text="Km", font=("Malgun Gothic", "20")).place(x=50, y = 355) #키로수
tkinter.Label(widget, textvariable=txt4, width=18, justify='center', bg='white', font=("Malgun Gothic", "20")).place(x=230, y = 355) # 게시용4

tkinter.Label(widget, text="No. of Owner", font=("Malgun Gothic", "20")).place(x=50, y = 400) #차량 소유자 수
tkinter.Label(widget, textvariable=txt5, width=18, justify='center', bg='white', font=("Malgun Gothic", "20")).place(x=230, y = 400) # 게시용5

tkinter.Label(widget, text="Start Price : RS.", font=("Malgun Gothic", "30")).place(x=50, y = 515) #차량 소유자 수
tkinter.Label(widget, textvariable=StartPrice, width=8, justify='center', font=("Malgun Gothic", "30")).place(x=350, y = 515) # 게시용6

tkinter.Label(widget, text="BID Price  : RS.", font=("Malgun Gothic", "30")).place(x=58, y = 590) #차량 소유자 수
Bidding = tkinter.Label(widget, textvariable=BidPrice, width=8, justify='center', fg = 'blue', font=("Malgun Gothic", "30")).place(x=350, y = 590) # 게시용7

tkinter.Label(widget, text="No.", font=("Malgun Gothic", "32")).place(x=150, y = 690)
tkinter.Label(widget, text="Applicant is", font=("Malgun Gothic", "20")).place(x=355, y = 705)
Bidders = tkinter.Label(widget, textvariable=BiderStatus, width=3, justify='center', fg = 'red', font=("Malgun Gothic", "45")).place(x=250, y = 670)
Bid_s = tkinter.Label(widget, textvariable=BidStatus, width=17, bg='white', justify='center', fg = 'red', font=("Malgun Gothic", "45")).place(x=10, y = 770)

#Nego = tkinter.Label(widget, textvariable=chk, justify='center', font=("Malgun Gothic", "25")).place(x=200, y = 700)

#tkinter.Label(widget, text="Squnece", font=30).place(x = 400, y = 250) #시퀀스
#tkinter.Label(widget, textvariable=txt1, width=30, bg='white', anchor='w').place(x = 350, y = 250) # 게시용1

#tkinter.Label(widget, text="Model", font=30).place(x = 400, y = 300) #모델명
#tkinter.Label(widget, textvariable=txt2, width=30, bg='white', anchor='w').place(x = 350, y = 300) # 게시용2

#tkinter.Label(widget, text="Regn No.", font=30).place(x = 400, y = 350) #등록번호
#tkinter.Label(widget, textvariable=txt3, width=30, bg='white', anchor='w').place(x = 350, y = 350) # 게시용3

#tkinter.Label(widget, text="Km", font=30).place(x = 400, y = 400) #키로수
#tkinter.Label(widget, textvariable=txt4, width=30, bg='white', anchor='w').place(x = 350, y = 400) # 게시용4

#tkinter.Label(widget, text="No. of Owner", font=30).place(x = 400, y = 450) #차량 소유자 수
#tkinter.Label(widget, textvariable=txt5, width=30, bg='white', anchor='w').place(x = 350, y = 450) # 게시용4


widget.mainloop()